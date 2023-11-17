#print ("hello,world!")
import openpyxl  


wb = openpyxl.load_workbook('C:\\Users\\Abil755rpp2\\Documents\\my git\\code\\python\\book1.xlsx')
ws = wb.active
print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
print('The value in cell A1 is: '+ws['B1'].value)
values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
print(values)
print(ws['A2'].value)